import http.server
import socketserver
import json
import os
import datetime
import threading
try:
    import cgi
except ImportError:
    cgi = None
import email.parser
import shutil

try:
    import openpyxl
except ImportError:
    openpyxl = None

PORT = int(os.environ.get("PORT", 8090))
WORKSPACE_DIR = os.path.dirname(os.path.abspath(__file__))
MAPPING_FILE = os.path.join(WORKSPACE_DIR, "uploads_mapping.json")

# In-memory thread-safe data cache
_cache = {"data": None, "timestamp": None}
_cache_lock = threading.Lock()

def parse_date(val):
    if not val or val in ("N/A", "-", "", "None"):
        return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        if isinstance(val, datetime.datetime):
            return val.date()
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%d", "%d-%b-%y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y"):
            try:
                return datetime.datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None

def format_date(dt):
    if dt is None:
        return "-"
    if isinstance(dt, (datetime.datetime, datetime.date)):
        return dt.strftime("%Y-%m-%d")
    return str(dt)

def load_uploads_mapping():
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[Warning] Error reading uploads_mapping.json: {e}")
    return {}

def save_uploads_mapping(mapping):
    try:
        with open(MAPPING_FILE, "w", encoding="utf-8") as f:
            json.dump(mapping, f, indent=2)
    except Exception as e:
        print(f"[Warning] Error saving uploads_mapping.json: {e}")

def find_active_files():
    mapping = load_uploads_mapping()
    files = os.listdir(WORKSPACE_DIR)
    
    result = {
        "wp1_topside_excl": None,
        "wp1_topside_structure": None,
        "wp1_jacket": None,
        "wp2_pipeline": None
    }
    
    # Check custom mapped files first
    for wp_key in result.keys():
        if wp_key in mapping and os.path.exists(os.path.join(WORKSPACE_DIR, mapping[wp_key])):
            result[wp_key] = os.path.join(WORKSPACE_DIR, mapping[wp_key])
            
    # Auto-detect remaining by keywords
    for fname in sorted(files, reverse=True):
        if not fname.endswith((".xlsx", ".xlsm")) or fname.startswith("~$"):
            continue
        fpath = os.path.join(WORKSPACE_DIR, fname)
        lower = fname.lower()
        if not result["wp1_topside_excl"] and ("00210" in lower or "excl" in lower):
            result["wp1_topside_excl"] = fpath
        elif not result["wp1_topside_structure"] and ("topsides" in lower or "topside structure" in lower or "topside_structure" in lower):
            result["wp1_topside_structure"] = fpath
        elif not result["wp1_jacket"] and ("jacket" in lower):
            result["wp1_jacket"] = fpath
        elif not result["wp2_pipeline"] and ("wp2" in lower or "pipeline" in lower):
            result["wp2_pipeline"] = fpath
            
    return result

def determine_doc_status(ifr_sub, ifa_sub, afc_sub):
    if afc_sub:
        return "AFC Submitted"
    if ifa_sub:
        return "IFA Submitted"
    if ifr_sub:
        return "IFR Submitted"
    return "Not Yet Submitted"

def extract_topside_excl(filepath):
    if not filepath or not os.path.exists(filepath):
        return {"docs": [], "filename": "Not Found"}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    del_map = {}
    del_sheet_name = "FilteredDEL" if "FilteredDEL" in wb.sheetnames else ("DEL" if "DEL" in wb.sheetnames else None)
    if del_sheet_name:
        ws = wb[del_sheet_name]
        empty_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                empty_count += 1
                if empty_count > 50:
                    break
                continue
            empty_count = 0
            del_id = str(row[0]).strip()
            title = str(row[1] or "Untitled Deliverable").strip() if len(row) > 1 else "Untitled Deliverable"
            discipline = str(row[4] or (row[6] if len(row) > 6 else None) or "General").strip() if len(row) > 4 else "General"
            if " - " in discipline:
                discipline = discipline.split(" - ")[0].strip()
            del_map[del_id] = {
                "doc_no": del_id,
                "title": title,
                "discipline": discipline,
                "ifr_plan": None, "ifr_forecast": None, "ifr_submit_date": None,
                "ifa_plan": None, "ifa_forecast": None, "ifa_submit_date": None,
                "afc_plan": None, "afc_forecast": None, "afc_submit_date": None
            }
            
    if "Gates" in wb.sheetnames:
        ws_g = wb["Gates"]
        empty_count = 0
        for row in ws_g.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                empty_count += 1
                if empty_count > 50:
                    break
                continue
            empty_count = 0
            del_id = str(row[0]).strip()
            if del_id not in del_map:
                continue
            status_id = str(row[2] or "").strip().upper() if len(row) > 2 else ""
            plan_dt = parse_date(row[10]) if len(row) > 10 else None
            actual_dt = parse_date(row[12]) if len(row) > 12 else None
            fore_dt = (parse_date(row[14]) if len(row) > 14 else None) or plan_dt
            
            entry = del_map[del_id]
            if status_id in ("IFR", "IFI", "IDC"):
                entry["ifr_plan"] = plan_dt
                entry["ifr_forecast"] = fore_dt
                entry["ifr_submit_date"] = actual_dt
            elif status_id in ("IFA",):
                entry["ifa_plan"] = plan_dt
                entry["ifa_forecast"] = fore_dt
                entry["ifa_submit_date"] = actual_dt
            elif status_id in ("AFC",):
                entry["afc_plan"] = plan_dt
                entry["afc_forecast"] = fore_dt
                entry["afc_submit_date"] = actual_dt
                
    wb.close()
    
    docs = []
    for d in del_map.values():
        d["status"] = determine_doc_status(d["ifr_submit_date"], d["ifa_submit_date"], d["afc_submit_date"])
        docs.append(d)
    return {"docs": docs, "filename": os.path.basename(filepath)}

def extract_topside_or_jacket(filepath, default_wp="Jacket"):
    if not filepath or not os.path.exists(filepath):
        return {"docs": [], "filename": "Not Found"}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if "EDSR" not in wb.sheetnames:
        wb.close()
        return {"docs": [], "filename": os.path.basename(filepath)}
        
    ws = wb["EDSR"]
    docs = []
    empty_count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or len(row) < 4 or not row[3]:
            empty_count += 1
            if empty_count > 50:
                break
            continue
        empty_count = 0
        doc_no = str(row[3]).strip()
        if doc_no in ("Total", "Average", "-", ""):
            continue
        title = str(row[4] or "Untitled Deliverable").strip() if len(row) > 4 else "Untitled Deliverable"
        discipline = str(row[1] or default_wp).strip() if len(row) > 1 else default_wp
        if discipline in ("GENERIC", "-", ""):
            discipline = default_wp
            
        ifr_plan = parse_date(row[10]) if len(row) > 10 else None
        ifr_fore = (parse_date(row[11]) if len(row) > 11 else None) or ifr_plan
        ifr_sub = parse_date(row[12]) if len(row) > 12 else None
        
        ifa_plan = parse_date(row[13]) if len(row) > 13 else None
        ifa_fore = (parse_date(row[14]) if len(row) > 14 else None) or ifa_plan
        ifa_sub = parse_date(row[15]) if len(row) > 15 else None
        
        afc_plan = (parse_date(row[16]) if len(row) > 16 else None) or (parse_date(row[19]) if len(row) > 19 else None)
        afc_fore = (parse_date(row[17]) if len(row) > 17 else None) or (parse_date(row[20]) if len(row) > 20 else None) or afc_plan
        afc_sub = (parse_date(row[18]) if len(row) > 18 else None) or (parse_date(row[21]) if len(row) > 21 else None)
        
        status = determine_doc_status(ifr_sub, ifa_sub, afc_sub)
        docs.append({
            "doc_no": doc_no,
            "title": title,
            "discipline": discipline,
            "ifr_plan": ifr_plan, "ifr_forecast": ifr_fore, "ifr_submit_date": ifr_sub,
            "ifa_plan": ifa_plan, "ifa_forecast": ifa_fore, "ifa_submit_date": ifa_sub,
            "afc_plan": afc_plan, "afc_forecast": afc_fore, "afc_submit_date": afc_sub,
            "status": status
        })
    wb.close()
    return {"docs": docs, "filename": os.path.basename(filepath)}

def extract_wp2_pipeline(filepath):
    if not filepath or not os.path.exists(filepath):
        return {"docs": [], "filename": "Not Found"}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if "MDR-Detailed Design-A10" not in wb.sheetnames:
        wb.close()
        return {"docs": [], "filename": os.path.basename(filepath)}
        
    ws = wb["MDR-Detailed Design-A10"]
    docs = []
    empty_count = 0
    for row in ws.iter_rows(min_row=11, values_only=True):
        if not row or len(row) < 9 or not row[8]:
            empty_count += 1
            if empty_count > 50:
                break
            continue
        empty_count = 0
        doc_no = str(row[8]).strip()
        if doc_no in ("Total", "Average", "-", ""):
            continue
        title = str(row[9] or "Untitled Deliverable").strip() if len(row) > 9 else "Untitled Deliverable"
        discipline = str(row[3] or row[2] or "Pipeline").strip() if len(row) > 3 else "Pipeline"
        if discipline in ("-", ""):
            discipline = "Pipeline"
            
        ifr_plan = parse_date(row[20]) if len(row) > 20 else None
        ifr_sub = parse_date(row[21]) if len(row) > 21 else None
        ifr_fore = (parse_date(row[22]) if len(row) > 22 else None) or ifr_plan
        
        ifa_plan = parse_date(row[29]) if len(row) > 29 else None
        ifa_sub = parse_date(row[30]) if len(row) > 30 else None
        ifa_fore = (parse_date(row[31]) if len(row) > 31 else None) or ifa_plan
        
        afc_plan = parse_date(row[36]) if len(row) > 36 else None
        afc_sub = parse_date(row[37]) if len(row) > 37 else None
        afc_fore = (parse_date(row[38]) if len(row) > 38 else None) or (parse_date(row[39]) if len(row) > 39 else None) or afc_plan
        
        status = determine_doc_status(ifr_sub, ifa_sub, afc_sub)
        docs.append({
            "doc_no": doc_no,
            "title": title,
            "discipline": discipline,
            "ifr_plan": ifr_plan, "ifr_forecast": ifr_fore, "ifr_submit_date": ifr_sub,
            "ifa_plan": ifa_plan, "ifa_forecast": ifa_fore, "ifa_submit_date": ifa_sub,
            "afc_plan": afc_plan, "afc_forecast": afc_fore, "afc_submit_date": afc_sub,
            "status": status
        })
    wb.close()
    return {"docs": docs, "filename": os.path.basename(filepath)}

def compute_delay_and_lookahead(docs, today):
    delayed_list = []
    lookahead_list = []
    
    delayed_type1 = 0 # Submitted Late
    delayed_type2 = 0 # Overdue & Not Submitted
    lookahead_slipping = 0 # Forecast > Plan
    on_time_docs_count = 0 # Deliverables without any delayed gates
    on_time_gates_count = 0 # Gates submitted on or before forecast date
    
    disc_summary = {}
    
    total_plan_prog = 0.0
    total_fore_prog = 0.0
    total_act_prog = 0.0
    
    for d in docs:
        disc = d["discipline"]
        if disc not in disc_summary:
            disc_summary[disc] = {
                "total": 0, "delayed": 0, "type3_1": 0, "type3_2": 0,
                "not_submitted": 0, "ifr_sub": 0, "ifa_sub": 0, "afc_sub": 0,
                "lookahead": 0, "act_prog_sum": 0.0, "plan_prog_sum": 0.0, "fore_prog_sum": 0.0
            }
        disc_summary[disc]["total"] += 1
        
        status = d["status"]
        if status == "AFC Submitted":
            disc_summary[disc]["afc_sub"] += 1
            act_prog = 100.0
        elif status == "IFA Submitted":
            disc_summary[disc]["ifa_sub"] += 1
            act_prog = 70.0
        elif status == "IFR Submitted":
            disc_summary[disc]["ifr_sub"] += 1
            act_prog = 50.0
        else:
            disc_summary[disc]["not_submitted"] += 1
            act_prog = 0.0
            
        if d["afc_plan"] is not None and d["afc_plan"] <= today:
            plan_prog = 100.0
        elif d["ifa_plan"] is not None and d["ifa_plan"] <= today:
            plan_prog = 70.0
        elif d["ifr_plan"] is not None and d["ifr_plan"] <= today:
            plan_prog = 50.0
        else:
            plan_prog = 0.0
            
        if d["afc_forecast"] is not None and d["afc_forecast"] <= today:
            fore_prog = 100.0
        elif d["ifa_forecast"] is not None and d["ifa_forecast"] <= today:
            fore_prog = 70.0
        elif d["ifr_forecast"] is not None and d["ifr_forecast"] <= today:
            fore_prog = 50.0
        else:
            fore_prog = 0.0
            
        total_act_prog += act_prog
        total_plan_prog += plan_prog
        total_fore_prog += fore_prog
        
        disc_summary[disc]["act_prog_sum"] += act_prog
        disc_summary[disc]["plan_prog_sum"] += plan_prog
        disc_summary[disc]["fore_prog_sum"] += fore_prog
            
        doc_delayed_entries = []
        doc_lookahead_entries = []
        
        gates = [
            ("IFR", d["ifr_plan"], d["ifr_forecast"], d["ifr_submit_date"]),
            ("IFA", d["ifa_plan"], d["ifa_forecast"], d["ifa_submit_date"]),
            ("AFC", d["afc_plan"], d["afc_forecast"], d["afc_submit_date"])
        ]
        
        # Identify the latest revision where actual date is empty (the pending revision)
        latest_pending_gate = next((g for g in gates if g[3] is None), None)
        
        for gate_name, plan_dt, fore_dt, sub_dt in gates:
            ref_dt = fore_dt if fore_dt is not None else plan_dt
            if ref_dt is None and sub_dt is None:
                continue
                
            if sub_dt is not None:
                if ref_dt is not None and sub_dt > ref_dt:
                    delay_days = max((sub_dt - ref_dt).days, 1)
                    doc_delayed_entries.append({
                        "doc_no": d["doc_no"],
                        "title": d["title"],
                        "discipline": d["discipline"],
                        "milestone": gate_name,
                        "delay_type": "Type 3.1 (Submitted Late)",
                        "delay_type_code": "3.1",
                        "delay_days": delay_days,
                        "plan_date": format_date(plan_dt),
                        "forecast_date": format_date(fore_dt),
                        "submit_date": format_date(sub_dt),
                        "urgency": "critical" if delay_days > 30 else ("high" if delay_days > 14 else "medium")
                    })
                    delayed_type1 += 1
                elif ref_dt is not None and sub_dt <= ref_dt:
                    on_time_gates_count += 1
            else:
                if ref_dt is not None:
                    if ref_dt <= today:
                        delay_days = max((today - ref_dt).days, 1)
                        doc_delayed_entries.append({
                            "doc_no": d["doc_no"],
                            "title": d["title"],
                            "discipline": d["discipline"],
                            "milestone": gate_name,
                            "delay_type": "Type 3.2 (Overdue & Not Submitted)",
                            "delay_type_code": "3.2",
                            "delay_days": delay_days,
                            "plan_date": format_date(plan_dt),
                            "forecast_date": format_date(fore_dt),
                            "submit_date": "-",
                            "urgency": "critical" if delay_days > 30 else ("high" if delay_days > 14 else "medium")
                        })
                        delayed_type2 += 1
                    elif latest_pending_gate and gate_name == latest_pending_gate[0] and (today < ref_dt <= today + datetime.timedelta(days=14)):
                        days_rem = (ref_dt - today).days
                        slipping = bool(plan_dt and fore_dt > plan_dt)
                        if slipping:
                            lookahead_slipping += 1
                        doc_lookahead_entries.append({
                            "doc_no": d["doc_no"],
                            "title": d["title"],
                            "discipline": d["discipline"],
                            "milestone": gate_name,
                            "days_remaining": days_rem,
                            "plan_date": format_date(plan_dt),
                            "forecast_date": format_date(fore_dt),
                            "slipping": slipping,
                            "urgency": "this_week" if days_rem <= 7 else "next_week"
                        })
                        
        if any(entry["delay_type_code"] == "3.1" for entry in doc_delayed_entries):
            disc_summary[disc]["type3_1"] += 1
        if any(entry["delay_type_code"] == "3.2" for entry in doc_delayed_entries):
            disc_summary[disc]["type3_2"] += 1
        if doc_lookahead_entries:
            disc_summary[disc]["lookahead"] += 1
                        
        if doc_delayed_entries:
            doc_delayed_entries.sort(key=lambda x: x["delay_days"], reverse=True)
            delayed_list.extend(doc_delayed_entries)
            disc_summary[disc]["delayed"] += 1
        else:
            on_time_docs_count += 1
            
        if doc_lookahead_entries:
            doc_lookahead_entries.sort(key=lambda x: x["days_remaining"])
            lookahead_list.extend(doc_lookahead_entries)
            
    serialized_docs = []
    for d in docs:
        sd = dict(d)
        for k in ("ifr_plan", "ifr_forecast", "ifr_submit_date",
                  "ifa_plan", "ifa_forecast", "ifa_submit_date",
                  "afc_plan", "afc_forecast", "afc_submit_date"):
            sd[k] = format_date(sd[k])
        serialized_docs.append(sd)
        
    total_docs = len(docs)
    plan_progress_pct = round(total_plan_prog / total_docs, 2) if total_docs > 0 else 0.0
    forecast_progress_pct = round(total_fore_prog / total_docs, 2) if total_docs > 0 else 0.0
    actual_progress_pct = round(total_act_prog / total_docs, 2) if total_docs > 0 else 0.0
    variance_pct = round(actual_progress_pct - plan_progress_pct, 2)
    spi = round(actual_progress_pct / plan_progress_pct, 2) if plan_progress_pct > 0 else (1.00 if actual_progress_pct >= 100.0 else 0.0)
    
    for disc, dsum in disc_summary.items():
        dtot = max(dsum["total"], 1)
        dsum["actual_progress_pct"] = round(dsum.get("act_prog_sum", 0.0) / dtot, 2)
        dsum["plan_progress_pct"] = round(dsum.get("plan_prog_sum", 0.0) / dtot, 2)
        dsum["forecast_progress_pct"] = round(dsum.get("fore_prog_sum", 0.0) / dtot, 2)
        
    return {
        "docs": serialized_docs,
        "delayed_list": delayed_list,
        "lookahead_list": lookahead_list,
        "kpi": {
            "total_docs": len(docs),
            "on_time_count": on_time_docs_count,
            "on_time_gates_count": on_time_gates_count,
            "delayed_count": len(delayed_list),
            "delayed_docs_count": len(set(d["doc_no"] for d in delayed_list)),
            "delayed_type1_count": delayed_type1,
            "delayed_type1_docs_count": len(set(d["doc_no"] for d in delayed_list if d["delay_type_code"] == "3.1")),
            "delayed_type2_count": delayed_type2,
            "delayed_type2_docs_count": len(set(d["doc_no"] for d in delayed_list if d["delay_type_code"] == "3.2")),
            "lookahead_count": len(lookahead_list),
            "lookahead_docs_count": len(set(d["doc_no"] for d in lookahead_list)),
            "lookahead_slipping_count": lookahead_slipping,
            "plan_progress_pct": plan_progress_pct,
            "forecast_progress_pct": forecast_progress_pct,
            "actual_progress_pct": actual_progress_pct,
            "variance_pct": variance_pct,
            "spi": spi
        },
        "discipline_summary": disc_summary
    }

def extract_official_summary_kpis(files):
    metrics = {
        "wp1_topside_excl": {"plan": 0.0, "forecast": 0.0, "actual": 0.0, "source": ""},
        "wp1_topside_structure": {"plan": 0.0, "forecast": 0.0, "actual": 0.0, "source": ""},
        "wp1_jacket": {"plan": 0.0, "forecast": 0.0, "actual": 0.0, "source": ""},
        "wp2_pipeline": {"plan": 0.0, "forecast": 0.0, "actual": 0.0, "source": "", "cutoff_date": "-"}
    }
    
    # 1. WP1 Topside Excl
    fpath = files.get("wp1_topside_excl")
    if fpath and os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "EDSR Report" in wb.sheetnames:
            ws = wb["EDSR Report"]
            for row in ws.iter_rows(min_row=2750, max_row=2810, values_only=True):
                if not row or len(row) < 34:
                    continue
                c1 = str(row[1] or "").strip().lower() if len(row) > 1 else ""
                c0 = str(row[0] or "").strip().lower() if len(row) > 0 else ""
                if c1 == "total :" or c0 == "total :" or c1 == "wk1 - work package 1 (greenfield) total :":
                    act_val = row[30] if len(row) > 30 and isinstance(row[30], (int, float)) else 0.0
                    plan_val = row[33] if len(row) > 33 and isinstance(row[33], (int, float)) else 0.0
                    if act_val > 0 or plan_val > 0:
                        metrics["wp1_topside_excl"]["actual"] = round(act_val * 100 if act_val <= 1.5 else act_val, 2)
                        metrics["wp1_topside_excl"]["plan"] = round(plan_val * 100 if plan_val <= 1.5 else plan_val, 2)
                        metrics["wp1_topside_excl"]["forecast"] = metrics["wp1_topside_excl"]["plan"]
                        metrics["wp1_topside_excl"]["source"] = "EDSR Report Row 2788 Total :"
                        if c1 == "total :" or c0 == "total :":
                            break
        wb.close()
        
    # 2. WP1 Topside Structure
    fpath = files.get("wp1_topside_structure")
    if fpath and os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "Progress Summary" in wb.sheetnames:
            ws = wb["Progress Summary"]
            for r_idx, row in enumerate(ws.iter_rows(min_row=7, max_row=12, values_only=True), 7):
                if not row or len(row) < 16:
                    continue
                title = " ".join([str(v) for v in row[:6] if v is not None]).strip().lower()
                if "overall topsides" in title or (r_idx == 8 and "topsides" in title):
                    plan_val = row[13] if isinstance(row[13], (int, float)) else 0.0
                    fore_val = row[14] if isinstance(row[14], (int, float)) else 0.0
                    act_val = row[15] if isinstance(row[15], (int, float)) else 0.0
                    metrics["wp1_topside_structure"]["plan"] = round(plan_val * 100 if plan_val <= 1.5 else plan_val, 2)
                    metrics["wp1_topside_structure"]["forecast"] = round(fore_val * 100 if fore_val <= 1.5 else fore_val, 2)
                    metrics["wp1_topside_structure"]["actual"] = round(act_val * 100 if act_val <= 1.5 else act_val, 2)
                    metrics["wp1_topside_structure"]["source"] = "Progress Summary Row Overall Topsides"
                    break
        wb.close()
        
    # 3. WP1 Jacket
    fpath = files.get("wp1_jacket")
    if fpath and os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "Progress Summary" in wb.sheetnames:
            ws = wb["Progress Summary"]
            for r_idx, row in enumerate(ws.iter_rows(min_row=11, max_row=15, values_only=True), 11):
                if not row or len(row) < 16:
                    continue
                title = " ".join([str(v) for v in row[:6] if v is not None]).strip().lower()
                if "overall jacket" in title or (r_idx == 12 and "jacket" in title):
                    plan_val = row[13] if isinstance(row[13], (int, float)) else 0.0
                    fore_val = row[14] if isinstance(row[14], (int, float)) else 0.0
                    act_val = row[15] if isinstance(row[15], (int, float)) else 0.0
                    metrics["wp1_jacket"]["plan"] = round(plan_val * 100 if plan_val <= 1.5 else plan_val, 2)
                    metrics["wp1_jacket"]["forecast"] = round(fore_val * 100 if fore_val <= 1.5 else fore_val, 2)
                    metrics["wp1_jacket"]["actual"] = round(act_val * 100 if act_val <= 1.5 else act_val, 2)
                    metrics["wp1_jacket"]["source"] = "Progress Summary Row Overall Jacket"
                    break
        wb.close()
        
    # 4. WP2 Pipeline
    fpath = files.get("wp2_pipeline")
    if fpath and os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        if "MDR-Detailed Design-A10" in wb.sheetnames:
            ws = wb["MDR-Detailed Design-A10"]
            plan_row = None
            actual_row = None
            fore_row = None
            months_row = None
            days_row = None
            
            for r_idx, row in enumerate(ws.iter_rows(min_row=7, max_row=225, values_only=True), 7):
                if not row:
                    continue
                if r_idx == 8:
                    months_row = row
                elif r_idx == 9:
                    days_row = row
                elif r_idx == 214:
                    plan_row = row
                elif r_idx == 216:
                    actual_row = row
                elif r_idx == 218:
                    fore_row = row
                    
            if actual_row and plan_row and fore_row:
                active_col = -1
                for col_idx in range(53, min(130, len(actual_row))):
                    val = actual_row[col_idx]
                    next_val = actual_row[col_idx + 1] if col_idx + 1 < len(actual_row) else None
                    if val is not None and isinstance(val, (int, float)) and val > 0:
                        if next_val is None or not isinstance(next_val, (int, float)) or next_val == 0:
                            active_col = col_idx
                            break
                if active_col != -1:
                    act_val = actual_row[active_col]
                    plan_val = plan_row[active_col] if len(plan_row) > active_col else 0.0
                    fore_val = fore_row[active_col] if len(fore_row) > active_col else 0.0
                    metrics["wp2_pipeline"]["actual"] = round(act_val * 100 if act_val <= 1.5 else act_val, 2)
                    metrics["wp2_pipeline"]["plan"] = round(plan_val * 100 if plan_val <= 1.5 else plan_val, 2)
                    metrics["wp2_pipeline"]["forecast"] = round(fore_val * 100 if fore_val <= 1.5 else fore_val, 2)
                    m_val = str(months_row[active_col] or "") if months_row and len(months_row) > active_col else ""
                    d_val = str(days_row[active_col] or "") if days_row and len(days_row) > active_col else ""
                    metrics["wp2_pipeline"]["cutoff_date"] = f"{d_val} {m_val}".strip()
                    metrics["wp2_pipeline"]["source"] = f"Row 214/216/218 at Col {active_col+1} ({metrics['wp2_pipeline']['cutoff_date']})"
        wb.close()
        
    return metrics

def extract_all_data(force_refresh=False):
    global _cache
    with _cache_lock:
        if not force_refresh and _cache["data"] is not None:
            return _cache["data"]
            
        today = datetime.date.today()
        files = find_active_files()
        
        print(f"[{datetime.datetime.now()}] Ingesting 4 Work Package Excel files...")
        wp1_topside_excl_raw = extract_topside_excl(files["wp1_topside_excl"])
        wp1_topside_str_raw = extract_topside_or_jacket(files["wp1_topside_structure"], default_wp="Topside Structure")
        wp1_jacket_raw = extract_topside_or_jacket(files["wp1_jacket"], default_wp="Jacket")
        wp2_pipeline_raw = extract_wp2_pipeline(files["wp2_pipeline"])
        
        wp1_topside_excl = compute_delay_and_lookahead(wp1_topside_excl_raw["docs"], today)
        wp1_topside_excl["filename"] = wp1_topside_excl_raw["filename"]
        
        wp1_topside_structure = compute_delay_and_lookahead(wp1_topside_str_raw["docs"], today)
        wp1_topside_structure["filename"] = wp1_topside_str_raw["filename"]
        
        wp1_jacket = compute_delay_and_lookahead(wp1_jacket_raw["docs"], today)
        wp1_jacket["filename"] = wp1_jacket_raw["filename"]
        
        wp2_pipeline = compute_delay_and_lookahead(wp2_pipeline_raw["docs"], today)
        wp2_pipeline["filename"] = wp2_pipeline_raw["filename"]
        
        all_docs = (
            wp1_topside_excl_raw["docs"] +
            wp1_topside_str_raw["docs"] +
            wp1_jacket_raw["docs"] +
            wp2_pipeline_raw["docs"]
        )
        executive = compute_delay_and_lookahead(all_docs, today)
        
        # Override KPIs with exact figures from official summary tables per user request
        print(f"[{datetime.datetime.now()}] Extracting official summary progress metrics...")
        official_metrics = extract_official_summary_kpis(files)
        
        wps_map = {
            "wp1_topside_excl": wp1_topside_excl,
            "wp1_topside_structure": wp1_topside_structure,
            "wp1_jacket": wp1_jacket,
            "wp2_pipeline": wp2_pipeline
        }
        
        total_docs_all = 0
        weighted_plan = 0.0
        weighted_forecast = 0.0
        weighted_actual = 0.0
        
        for wp_key, wp_obj in wps_map.items():
            off = official_metrics.get(wp_key)
            if off and (off["plan"] > 0 or off["actual"] > 0):
                wp_obj["kpi"]["plan_progress_pct"] = off["plan"]
                wp_obj["kpi"]["forecast_progress_pct"] = off["forecast"]
                wp_obj["kpi"]["actual_progress_pct"] = off["actual"]
                wp_obj["kpi"]["variance_pct"] = round(off["actual"] - off["plan"], 2)
                wp_obj["kpi"]["spi"] = round(off["actual"] / off["plan"], 2) if off["plan"] > 0 else 0.0
                if wp_key == "wp2_pipeline" and off.get("cutoff_date", "-") != "-":
                    wp_obj["kpi"]["cutoff_date"] = off["cutoff_date"]
            
            docs_cnt = max(1, wp_obj["kpi"].get("total_docs", 1))
            total_docs_all += docs_cnt
            weighted_plan += wp_obj["kpi"]["plan_progress_pct"] * docs_cnt
            weighted_forecast += wp_obj["kpi"]["forecast_progress_pct"] * docs_cnt
            weighted_actual += wp_obj["kpi"]["actual_progress_pct"] * docs_cnt
            
        if total_docs_all > 0:
            executive["kpi"]["plan_progress_pct"] = round(weighted_plan / total_docs_all, 2)
            executive["kpi"]["forecast_progress_pct"] = round(weighted_forecast / total_docs_all, 2)
            executive["kpi"]["actual_progress_pct"] = round(weighted_actual / total_docs_all, 2)
            executive["kpi"]["variance_pct"] = round(executive["kpi"]["actual_progress_pct"] - executive["kpi"]["plan_progress_pct"], 2)
            executive["kpi"]["spi"] = round(executive["kpi"]["actual_progress_pct"] / executive["kpi"]["plan_progress_pct"], 2) if executive["kpi"]["plan_progress_pct"] > 0 else 0.0

        executive["filenames"] = {
            "Topside Excl. Structure": wp1_topside_excl["filename"],
            "Topside Structure": wp1_topside_structure["filename"],
            "Jacket": wp1_jacket["filename"],
            "Pipeline": wp2_pipeline["filename"]
        }
        executive["wp_summary"] = [
            {
                "wp_key": "wp1_topside_excl",
                "wp_name": "WP-1 Topside Excl. Structure",
                "filename": wp1_topside_excl["filename"],
                "kpi": wp1_topside_excl["kpi"]
            },
            {
                "wp_key": "wp1_topside_structure",
                "wp_name": "WP-1 Topside Structure",
                "filename": wp1_topside_structure["filename"],
                "kpi": wp1_topside_structure["kpi"]
            },
            {
                "wp_key": "wp1_jacket",
                "wp_name": "WP-1 Jacket",
                "filename": wp1_jacket["filename"],
                "kpi": wp1_jacket["kpi"]
            },
            {
                "wp_key": "wp2_pipeline",
                "wp_name": "WP-2 Pipeline",
                "filename": wp2_pipeline["filename"],
                "kpi": wp2_pipeline["kpi"]
            }
        ]
        
        payload = {
            "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "today": today.strftime("%Y-%m-%d"),
            "executive": executive,
            "wp1_topside_excl": wp1_topside_excl,
            "wp1_topside_structure": wp1_topside_structure,
            "wp1_jacket": wp1_jacket,
            "wp2_pipeline": wp2_pipeline
        }
        
        _cache["data"] = payload
        _cache["timestamp"] = datetime.datetime.now()
        return payload

class DashboardRequestHandler(http.server.SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=WORKSPACE_DIR, **kwargs)
        
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        
    def do_GET(self):
        if self.path.startswith("/api/data"):
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            data = extract_all_data()
            self.wfile.write(json.dumps(data).encode("utf-8"))
        elif self.path.startswith("/api/status"):
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            files = find_active_files()
            status = {
                "server_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "files_detected": {k: os.path.basename(v) if v else "None" for k, v in files.items()}
            }
            self.wfile.write(json.dumps(status).encode("utf-8"))
        else:
            super().do_GET()
            
    def do_POST(self):
        if self.path.startswith("/api/upload"):
            wp_key = ""
            filename = ""
            save_path = ""
            
            if cgi is not None:
                form = cgi.FieldStorage(
                    fp=self.rfile,
                    headers=self.headers,
                    environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers["Content-Type"]}
                )
                
                wp_key = form.getvalue("wp_key", "").strip()
                if not wp_key or wp_key not in ("wp1_topside_excl", "wp1_topside_structure", "wp1_jacket", "wp2_pipeline"):
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Invalid or missing wp_key parameter"}).encode("utf-8"))
                    return
                    
                if "file" not in form:
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "No file uploaded in 'file' field"}).encode("utf-8"))
                    return
                    
                file_item = form["file"]
                if not file_item.filename:
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Empty filename"}).encode("utf-8"))
                    return
                    
                filename = os.path.basename(file_item.filename)
                save_path = os.path.join(WORKSPACE_DIR, filename)
                
                with open(save_path, "wb") as f:
                    while True:
                        chunk = file_item.file.read(65536)
                        if not chunk:
                            break
                        f.write(chunk)
            else:
                content_length = int(self.headers.get("Content-Length", 0))
                if content_length <= 0:
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Empty request body"}).encode("utf-8"))
                    return
                body = self.rfile.read(content_length)
                content_type = self.headers.get("Content-Type", "")
                header_bytes = f"Content-Type: {content_type}\r\n\r\n".encode("latin1") + body
                msg = email.parser.BytesParser().parsebytes(header_bytes)
                
                if not msg.is_multipart():
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Expected multipart/form-data"}).encode("utf-8"))
                    return
                    
                file_data = None
                for part in msg.get_payload():
                    name = part.get_param("name", header="content-disposition")
                    part_filename = part.get_filename()
                    if name == "wp_key":
                        wp_key = part.get_payload(decode=True).decode("utf-8", errors="ignore").strip()
                    elif name == "file" or part_filename:
                        if part_filename:
                            filename = os.path.basename(part_filename)
                            file_data = part.get_payload(decode=True)
                            
                if not wp_key or wp_key not in ("wp1_topside_excl", "wp1_topside_structure", "wp1_jacket", "wp2_pipeline"):
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "Invalid or missing wp_key parameter"}).encode("utf-8"))
                    return
                if not filename or file_data is None:
                    self.send_response(400)
                    self.send_header("Content-Type", "application/json")
                    self.end_headers()
                    self.wfile.write(json.dumps({"error": "No file uploaded in 'file' field"}).encode("utf-8"))
                    return
                save_path = os.path.join(WORKSPACE_DIR, filename)
                with open(save_path, "wb") as f:
                    f.write(file_data)
                    
            mapping = load_uploads_mapping()
            mapping[wp_key] = filename
            save_uploads_mapping(mapping)
            
            global _cache
            with _cache_lock:
                _cache["data"] = None
            extract_all_data(force_refresh=True)
            
            self.send_response(200)
            self.send_header("Content-Type", "application/json; charset=utf-8")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({
                "status": "success",
                "wp_key": wp_key,
                "filename": filename,
                "message": f"Successfully uploaded and mapped {filename} to {wp_key}"
            }).encode("utf-8"))
        else:
            self.send_response(404)
            self.end_headers()

class ThreadedHTTPServer(socketserver.ThreadingMixIn, http.server.HTTPServer):
    daemon_threads = True

if __name__ == "__main__":
    if openpyxl is None:
        print("[Error] openpyxl library is required. Please install it using `pip install openpyxl`.")
        exit(1)
        
    print("=== Initializing Z1F Engineering Progress Dashboard Server ===")
    t0 = datetime.datetime.now()
    data = extract_all_data(force_refresh=True)
    t1 = datetime.datetime.now()
    kpi = data["executive"]["kpi"]
    print(f"[Ready in {(t1-t0).total_seconds():.2f}s] Total Docs: {kpi['total_docs']} | Total Delayed: {kpi['delayed_count']} (Type 3.1: {kpi['delayed_type1_count']} | Type 3.2: {kpi['delayed_type2_count']}) | Lookahead Risk: {kpi['lookahead_slipping_count']}")
    
    server = ThreadedHTTPServer(("", PORT), DashboardRequestHandler)
    print(f"[Server Running] Dashboard live at: http://localhost:{PORT}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n[Shutdown] Server stopped gracefully.")
        server.server_close()
